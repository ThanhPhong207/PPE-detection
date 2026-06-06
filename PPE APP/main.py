import cv2
import os
import sys
import time
import subprocess
import threading
import tkinter as tk
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk
from ultralytics import YOLO
from datetime import datetime
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
import io

# Import class từ file image_processor.py
from image_processor import ImageEnhancer

# --- HẰNG SỐ & TIỆN ÍCH KIỂM TRA PPE THEO TỪNG NGƯỜI ---
CONTAINMENT_THRESHOLD = 0.5   # tỉ lệ box đồ bảo hộ phải nằm trong box người để được gán cho người đó
DEBOUNCE_SECONDS = 2.0        # thời gian xác nhận vi phạm (chống nhấp nháy)
RESET_GRACE = 0.5             # giây: scene phải sạch liên tục lâu hơn ngần này mới xoá bộ đếm
                              # -> dropout/nhận nhầm 1-2 frame không làm mất thời gian đã tích luỹ

# Ánh xạ item trên UI -> (class khẳng định, class phủ định) của model
ITEM_MAP = {
    'Hardhat': ('Hardhat', 'NO-Hardhat'),
    'Safety Vest': ('Safety Vest', 'NO-Safety Vest'),
    'Mask': ('Mask', 'NO-Mask'),
}
GEAR_NAMES = {'Hardhat', 'Mask', 'Safety Vest', 'NO-Hardhat', 'NO-Mask', 'NO-Safety Vest'}

def _area(b):
    return max(0.0, b[2] - b[0]) * max(0.0, b[3] - b[1])

def _inter(a, b):
    x1, y1 = max(a[0], b[0]), max(a[1], b[1])
    x2, y2 = min(a[2], b[2]), min(a[3], b[3])
    return max(0.0, x2 - x1) * max(0.0, y2 - y1)

def _center(b):
    return ((b[0] + b[2]) / 2.0, (b[1] + b[3]) / 2.0)

# --- QUẢN LÝ EXCEL ---
class ViolationLogger:
    def __init__(self, folder="violations"):
        self.folder = folder
        if not os.path.exists(self.folder):
            os.makedirs(self.folder)
            
        self.start_time_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.filename = os.path.join(self.folder, f"baocao_PPE_{self.start_time_str}.xlsx")
        self.columns = ['Thời gian', 'Loại vi phạm', 'Lý do', 'Hình ảnh']
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vi phạm"
        ws.append(self.columns)
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 30
        wb.save(self.filename)

    def log_to_excel(self, violations, frame):
        try:
            wb = openpyxl.load_workbook(self.filename)
            ws = wb.active
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            violation_str = ", ".join(violations)
            
            current_row = ws.max_row + 1
            ws.append([now, violation_str, f"Phát hiện thiếu: {violation_str.upper()}", ""]) 

            success, encoded_image = cv2.imencode('.jpg', frame)
            if success:
                img_data = io.BytesIO(encoded_image.tobytes())
                img = OpenpyxlImage(img_data)
                img.width, img.height = 200, 150
                ws.row_dimensions[current_row].height = 120 
                ws.add_image(img, f'D{current_row}')

            wb.save(self.filename)
        except Exception as e:
            print(f"Lỗi ghi Excel: {e}")

# --- BỘ NHẬN DIỆN YOLO: KIỂM TRA PPE ĐỘC LẬP TỪNG NGƯỜI ---
class PPEDetector:
    def __init__(self, model_path):
        self.model = YOLO(model_path)
        self.class_names = self.model.names
        self.violation_start = None   # thời điểm bắt đầu chuỗi frame liên tục CÓ vi phạm
        self.clear_since = None       # thời điểm scene bắt đầu sạch (để áp dụng ân hạn reset)

    def detect_and_check(self, frame, required_items, still=False):
        # Tiền xử lý: chỉ giữ CLAHE (bỏ blur + sharpen vì làm méo đặc trưng YOLO)
        proc = ImageEnhancer.apply_clahe(frame)
        results = self.model(proc, stream=True, conf=0.4)

        persons, gears = [], []
        for r in results:
            for box in r.boxes:
                name = self.class_names.get(int(box.cls), "Unknown")
                xyxy = box.xyxy[0].tolist()
                conf = float(box.conf)
                if name == 'Person':
                    persons.append({'box': xyxy, 'gear': []})
                elif name in GEAR_NAMES:
                    gears.append({'name': name, 'conf': conf, 'box': xyxy, 'c': _center(xyxy)})

        # Gán mỗi box đồ bảo hộ cho đúng MỘT người theo độ chứa (containment), conf cao xét trước
        for g in sorted(gears, key=lambda g: -g['conf']):
            cand = [(_inter(g['box'], p['box']) / max(_area(g['box']), 1e-6), p) for p in persons]
            if not cand:
                continue
            max_c = max(c for c, _ in cand)
            if max_c < CONTAINMENT_THRESHOLD:
                continue   # đồ bảo hộ không nằm đủ trong ai -> bỏ qua
            # Trong các ứng viên gần hoà (chênh <=0.05), chọn người có tâm box gần đồ bảo hộ nhất
            gc = g['c']
            near = [p for c, p in cand if max_c - c <= 0.05]
            best_p = min(near, key=lambda p: (gc[0]-_center(p['box'])[0])**2 + (gc[1]-_center(p['box'])[1])**2)
            best_p['gear'].append(g)

        # --- Kiểm tra KHÔNG GIAN: từng người độc lập trong frame này ---
        per_person = []        # [(box, missing_items)]
        frame_missing = set()
        any_violating = False
        for p in persons:
            missing = []
            for item in required_items:
                pos_name, neg_name = ITEM_MAP[item]
                pos = max([g['conf'] for g in p['gear'] if g['name'] == pos_name], default=None)
                neg = max([g['conf'] for g in p['gear'] if g['name'] == neg_name], default=None)
                if neg is not None and (pos is None or neg >= pos):
                    missing.append(item)            # NO-* là tín hiệu vi phạm model đã học -> thiếu
                elif pos is not None and (neg is None or pos > neg):
                    pass                            # có đồ bảo hộ -> đạt
                else:
                    missing.append(item)            # không tín hiệu nào -> coi như thiếu (fail-closed)
            per_person.append((p['box'], missing))
            if missing:
                any_violating = True
                frame_missing.update(missing)

        # --- Xác nhận theo THỜI GIAN: một bộ đếm chung ---
        # Chỉ cần frame liên tục CÓ vi phạm đủ lâu là báo động; không bám danh tính từng người
        # (tránh hồi quy false-SAFE khi người di chuyển/đứng cạnh nhau). Có ân hạn reset để
        # detector rớt box 1-2 frame không xoá mất thời gian đã tích luỹ.
        now = time.time()
        if still:
            confirmed, remaining = any_violating, 0
        else:
            if any_violating:
                self.clear_since = None
                if self.violation_start is None:
                    self.violation_start = now
            elif self.violation_start is not None:
                if self.clear_since is None:
                    self.clear_since = now
                if now - self.clear_since >= RESET_GRACE:
                    self.violation_start = self.clear_since = None
            if self.violation_start is not None:
                elapsed = now - self.violation_start
                confirmed = elapsed >= DEBOUNCE_SECONDS
                remaining = max(1, int(DEBOUNCE_SECONDS - elapsed))
            else:
                confirmed, remaining = False, 0

        # --- Vẽ: mỗi người tô màu theo trạng thái không gian của họ + cờ confirmed chung ---
        annotated = frame.copy()
        GREEN, YELLOW, RED, CYAN = (0, 255, 0), (0, 255, 255), (0, 0, 255), (255, 255, 0)
        FONT = cv2.FONT_HERSHEY_SIMPLEX
        for box, missing in per_person:
            x1, y1, x2, y2 = [int(v) for v in box]
            if not missing:
                col, lbl = GREEN, "OK"
            elif confirmed:
                col, lbl = RED, "MISSING: " + ", ".join(missing)
            else:
                col, lbl = YELLOW, "VERIFYING (%ds)" % remaining
            cv2.rectangle(annotated, (x1, y1), (x2, y2), col, 2)
            cv2.putText(annotated, lbl, (x1, max(20, y1 - 8)), FONT, 0.6, col, 2)

        # Banner tổng hợp suy ra từ cùng cờ confirmed -> luôn khớp màu box
        if not persons:
            is_safe, banner, bcol = True, "SCANNING...", CYAN
        elif confirmed:
            is_safe, banner, bcol = False, "DANGER: MISSING " + ", ".join(sorted(frame_missing)), RED
        elif any_violating:
            is_safe, banner, bcol = True, "VERIFYING...", YELLOW
        else:
            is_safe, banner, bcol = True, "STATUS: SAFE", GREEN

        cv2.putText(annotated, banner, (20, 50), FONT, 0.8, bcol, 2)
        return annotated, is_safe, sorted(frame_missing)

# --- GIAO DIỆN NGƯỜI DÙNG ---
class PPEApp:
    def __init__(self, window):
        self.window = window
        self.window.title("GIÁM SÁT PPE - ENHANCED VERSION")
        self.window.geometry("1050x700")
        
        self.excel_logger = ViolationLogger()
        self.detector = PPEDetector("weights/best.pt") # Đảm bảo file weight đúng đường dẫn
        self.is_running, self.last_log_time = False, 0
        self._init_ui()

    def _init_ui(self):
        tk.Label(self.window, text="HỆ THỐNG GIÁM SÁT AN TOÀN PPE", font=('Arial', 20, 'bold'), bg='#2c3e50', fg='white', pady=10).pack(fill=tk.X)
        container = tk.Frame(self.window, bg='#ecf0f1')
        container.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)

        self.canvas = tk.Canvas(container, width=800, height=500, bg="black")
        self.canvas.pack(side=tk.LEFT, padx=10)

        right_panel = tk.Frame(container, bg='#ecf0f1')
        right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        tk.Label(right_panel, text="CÀI ĐẶT KIỂM TRA", font=('Arial', 12, 'bold'), bg='#ecf0f1').pack(pady=10)
        self.check_vars = {k: tk.BooleanVar(value=True) for k in ['Hardhat', 'Safety Vest', 'Mask']}
        for text, var in self.check_vars.items():
            tk.Checkbutton(right_panel, text=text, variable=var, bg='#ecf0f1', font=('Arial', 11)).pack(anchor=tk.W, padx=20)

        btn_s = {'width': 20, 'font': ('Arial', 10, 'bold'), 'pady': 5}
        tk.Button(right_panel, text="📁 CHỌN ẢNH", command=self.process_image, **btn_s).pack(pady=5)
        tk.Button(right_panel, text="🎬 CHỌN VIDEO", command=self.process_video, **btn_s).pack(pady=5)
        tk.Button(right_panel, text="📊 MỞ FILE EXCEL", command=self.open_excel, bg='#27ae60', fg='white', **btn_s).pack(pady=5)
        tk.Button(right_panel, text="🛑 DỪNG", command=self.stop_stream, bg='#e74c3c', fg='white', **btn_s).pack(pady=15)

        self.status_lbl = tk.Label(self.window, text="SẴN SÀNG", font=('Arial', 14, 'bold'), bg='#34495e', fg='white', pady=5)
        self.status_lbl.pack(fill=tk.X, side=tk.BOTTOM)

    def update_display(self, frame):
        img_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        img_pil = Image.fromarray(img_rgb)
        img_pil.thumbnail((800, 500))
        self.photo = ImageTk.PhotoImage(image=img_pil)
        self.canvas.create_image(400, 250, image=self.photo)

    def process_image(self):
        path = filedialog.askopenfilename()
        if path: 
            self.stop_stream()
            frame = cv2.imread(path)
            if frame is not None:
                self._run_logic(frame, still=True)

    def process_video(self):
        path = filedialog.askopenfilename()
        if path:
            self.stop_stream()
            self.detector.violation_start = self.detector.clear_since = None   # tránh kế thừa bộ đếm cũ
            self.is_running = True
            threading.Thread(target=self.video_loop, args=(path,), daemon=True).start()

    def video_loop(self, path):
        cap = cv2.VideoCapture(path)
        while self.is_running and cap.isOpened():
            ret, frame = cap.read()
            if not ret: break
            self._run_logic(frame)
            time.sleep(0.01)
        cap.release()
        self.is_running = False

    def _run_logic(self, frame, still=False):
        req = [k for k, v in self.check_vars.items() if v.get()]
        res_f, safe, viol = self.detector.detect_and_check(frame, req, still=still)
        self.window.after(0, self.update_display, res_f)
        
        if not safe:
            self.window.after(0, lambda: self.status_lbl.config(text=f"CẢNH BÁO: THIẾU {', '.join(viol).upper()}", bg='#e74c3c'))
            if time.time() - self.last_log_time > 5:
                self.excel_logger.log_to_excel(viol, frame)
                self.last_log_time = time.time()
        else:
            self.window.after(0, lambda: self.status_lbl.config(text="TRẠNG THÁI: AN TOÀN", bg='#27ae60'))

    def open_excel(self):
        path = os.path.abspath(self.excel_logger.filename)
        try:
            if sys.platform.startswith('win'):
                os.startfile(path)
            elif sys.platform == 'darwin':
                subprocess.run(['open', path])
            else:
                subprocess.run(['xdg-open', path])
        except Exception:
            messagebox.showerror("Lỗi", "Không thể mở file.")

    def stop_stream(self):
        self.is_running = False
        time.sleep(0.2)
        self.canvas.delete("all")
        self.status_lbl.config(text="SẴN SÀNG", bg='#34495e')

if __name__ == "__main__":
    root = tk.Tk()
    PPEApp(root)
    root.mainloop()